"""Exporter les datasets KG/catalogues dans un classeur de comparaison.

Le classeur aide a comparer les nœuds Dataset du KG avec les catalogues
software R/Python et a reperer les doublons ou les entrees peu utiles.
"""

from __future__ import annotations

import csv
import json
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DATASETS_DIR = ROOT / "data" / "manifests" / "datasets"
GRAPH = ROOT / ".kg" / "graph.sqlite"
OUT = DATASETS_DIR / f"dataset_kg_catalog_comparison_{date.today().isoformat().replace('-', '_')}.xlsx"


def normalise(value: object) -> str:
    """Normaliser un libelle pour faciliter la detection des doublons."""
    text = "" if value is None else str(value).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    """Lire un CSV en conservant les noms de colonnes existants."""
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def first_present(row: dict[str, object], *names: str) -> str:
    """Retourner la premiere colonne non vide parmi plusieurs noms possibles."""
    for name in names:
        value = row.get(name)
        if value not in (None, ""):
            return str(value)
    return ""


def jsonable_row(row: dict[object, object]) -> dict[str, object]:
    """Convertir les cles CSV eventuellement vides en chaines JSON valides."""
    return {"" if key is None else str(key): value for key, value in row.items()}


def count_edges(conn: sqlite3.Connection, dataset_id: str, relation: str) -> int:
    """Compter les relations sortantes d'un dataset dans le KG."""
    return conn.execute(
        "select count(*) from edges where source=? and relation=?",
        (dataset_id, relation),
    ).fetchone()[0]


def collect_kg_datasets() -> list[dict[str, object]]:
    """Extraire tous les nœuds Dataset du graphe SQLite."""
    conn = sqlite3.connect(GRAPH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "select id, label, props_json from nodes where type='Dataset' order by id"
    ).fetchall()

    out: list[dict[str, object]] = []
    for row in rows:
        props = json.loads(row["props_json"] or "{}")
        package = first_present(props, "package", "pkg")
        dataset = first_present(props, "dataset", "name")
        source = first_present(props, "source")
        label = row["label"]
        norm_key = normalise(f"{package}::{dataset}" if package or dataset else label)
        variable_count = count_edges(conn, row["id"], "HAS_VARIABLE")
        response_count = count_edges(conn, row["id"], "HAS_RESPONSE")
        covariate_count = count_edges(conn, row["id"], "HAS_COVARIATE")
        formula_count = count_edges(conn, row["id"], "SHOWS_FORMULA")
        paper_count = conn.execute(
            "select count(*) from edges where target=? and relation='USES_DATASET'",
            (row["id"],),
        ).fetchone()[0]
        documentation_count = count_edges(conn, row["id"], "DOCUMENTED_BY")

        weak_flags: list[str] = []
        if source in {"", "r_dataset_docs"}:
            weak_flags.append("source_doc_or_missing")
        if variable_count <= 2:
            weak_flags.append("few_variables")
        if response_count == 0 and formula_count == 0:
            weak_flags.append("no_response_or_formula")
        if re.search(r"(geometry|geometries|region|boundary|shape|shapefile|map|grid|mask|points|polygons?)", label, re.I):
            weak_flags.append("geometry_like")

        out.append(
            {
                "kg_id": row["id"],
                "label": label,
                "package": package,
                "dataset": dataset,
                "source": source or "<missing>",
                "normalised_key": norm_key,
                "variable_count": variable_count,
                "response_count": response_count,
                "covariate_count": covariate_count,
                "formula_count": formula_count,
                "paper_count": paper_count,
                "documentation_count": documentation_count,
                "review_flags": "; ".join(weak_flags),
                "props_json": json.dumps(props, ensure_ascii=False, sort_keys=True),
            }
        )
    return out


def collect_catalog_rows() -> list[dict[str, object]]:
    """Lire les catalogues software et produire une table commune."""
    specs = [
        ("software_r_catalog_main", DATASETS_DIR / "software_r_catalog_main_datasets.csv"),
        ("software_python_catalog_all", DATASETS_DIR / "software_python_catalog_all.csv"),
        ("software_catalog_curated_final", DATASETS_DIR / "software_catalog_curated_final.csv"),
        ("software_r_paper_formula_audit", DATASETS_DIR / "software_r_dataset_paper_formula_audit.csv"),
    ]

    rows: list[dict[str, object]] = []
    for source_name, path in specs:
        for row in read_csv_rows(path):
            package = first_present(row, "package", "pkg", "Package")
            dataset = first_present(row, "dataset", "name", "Dataset")
            label = first_present(row, "label", "title", "description", "dataset", "Dataset")
            norm_key = normalise(f"{package}::{dataset}" if package or dataset else label)
            rows.append(
                {
                    "catalog_source": source_name,
                    "package": package,
                    "dataset": dataset,
                    "label": label,
                    "normalised_key": norm_key,
                    "is_spatial": first_present(row, "is_spatial", "spatial", "has_geometry"),
                    "is_spatiotemporal": first_present(row, "is_spatiotemporal", "spatiotemporal", "has_time"),
                    "n_rows": first_present(row, "n_rows", "rows", "nrow"),
                    "n_cols": first_present(row, "n_cols", "columns", "ncol"),
                    "response": first_present(row, "response", "response_variable", "y", "candidate_y"),
                    "covariates": first_present(row, "covariates", "candidate_x", "x", "predictors"),
                    "formula": first_present(row, "formula", "model_formula"),
                    "raw_row_json": json.dumps(jsonable_row(row), ensure_ascii=False, sort_keys=True),
                }
            )
    return rows


def write_sheet(ws, rows: list[dict[str, object]], headers: list[str]) -> None:
    """Ecrire un onglet avec filtres, gel et largeurs lisibles."""
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, header in enumerate(headers, start=1):
        width = min(max(len(header) + 2, 12), 60)
        for value in ws.iter_cols(min_col=idx, max_col=idx, min_row=2, max_row=min(ws.max_row, 200), values_only=True):
            for item in value:
                width = min(max(width, len(str(item)) + 2), 70)
        ws.column_dimensions[get_column_letter(idx)].width = width


def main() -> None:
    kg_rows = collect_kg_datasets()
    catalog_rows = collect_catalog_rows()

    kg_key_counts = Counter(row["normalised_key"] for row in kg_rows)
    catalog_key_counts = Counter(row["normalised_key"] for row in catalog_rows)

    for row in kg_rows:
        row["kg_duplicate_count"] = kg_key_counts[row["normalised_key"]]
        row["catalog_match_count"] = catalog_key_counts.get(row["normalised_key"], 0)

    for row in catalog_rows:
        row["catalog_duplicate_count"] = catalog_key_counts[row["normalised_key"]]
        row["kg_match_count"] = kg_key_counts.get(row["normalised_key"], 0)

    duplicate_rows = [
        {
            "normalised_key": key,
            "kg_count": kg_key_counts.get(key, 0),
            "catalog_count": catalog_key_counts.get(key, 0),
            "example_labels": " | ".join(
                sorted(
                    {
                        str(row.get("label", ""))
                        for row in kg_rows + catalog_rows
                        if row.get("normalised_key") == key
                    }
                )[:5]
            ),
        }
        for key in sorted(set(kg_key_counts) | set(catalog_key_counts))
        if kg_key_counts.get(key, 0) > 1
        or catalog_key_counts.get(key, 0) > 1
        or (kg_key_counts.get(key, 0) and catalog_key_counts.get(key, 0))
    ]

    weak_rows = [
        row
        for row in kg_rows
        if row["review_flags"]
        or int(row["variable_count"]) <= 2
        or int(row["response_count"]) == 0
        and int(row["formula_count"]) == 0
    ]

    summary_rows = [
        {"metric": "KG Dataset nodes", "value": len(kg_rows)},
        {"metric": "Catalog rows combined", "value": len(catalog_rows)},
        {"metric": "Unique KG normalised keys", "value": len(kg_key_counts)},
        {"metric": "Unique catalog normalised keys", "value": len(catalog_key_counts)},
        {"metric": "Duplicate/cross-match keys", "value": len(duplicate_rows)},
        {"metric": "KG rows flagged for review", "value": len(weak_rows)},
    ]
    for source, count in Counter(row["source"] for row in kg_rows).most_common():
        summary_rows.append({"metric": f"KG source: {source}", "value": count})
    for source, count in Counter(row["catalog_source"] for row in catalog_rows).most_common():
        summary_rows.append({"metric": f"Catalog source: {source}", "value": count})

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    write_sheet(ws, summary_rows, ["metric", "value"])

    ws = wb.create_sheet("KG_Datasets")
    write_sheet(
        ws,
        kg_rows,
        [
            "kg_id",
            "label",
            "package",
            "dataset",
            "source",
            "normalised_key",
            "kg_duplicate_count",
            "catalog_match_count",
            "variable_count",
            "response_count",
            "covariate_count",
            "formula_count",
            "paper_count",
            "documentation_count",
            "review_flags",
            "props_json",
        ],
    )

    ws = wb.create_sheet("Software_Catalogs")
    write_sheet(
        ws,
        catalog_rows,
        [
            "catalog_source",
            "package",
            "dataset",
            "label",
            "normalised_key",
            "catalog_duplicate_count",
            "kg_match_count",
            "is_spatial",
            "is_spatiotemporal",
            "n_rows",
            "n_cols",
            "response",
            "covariates",
            "formula",
            "raw_row_json",
        ],
    )

    ws = wb.create_sheet("Duplicate_Candidates")
    write_sheet(ws, duplicate_rows, ["normalised_key", "kg_count", "catalog_count", "example_labels"])

    ws = wb.create_sheet("Review_Low_Utility")
    write_sheet(
        ws,
        weak_rows,
        [
            "kg_id",
            "label",
            "package",
            "dataset",
            "source",
            "normalised_key",
            "variable_count",
            "response_count",
            "covariate_count",
            "formula_count",
            "paper_count",
            "documentation_count",
            "review_flags",
        ],
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
